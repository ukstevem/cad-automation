"""
Generate an Excel (.xlsx) BOM workbook with embedded STL thumbnails.

Reads the sidecar JSON to gather CNC results, classifications, STL map,
and consolidation groups, then builds a multi-sheet workbook:
  - CNC Items        (sections + plates with full detail)
  - Unknown Items    (postprocess but type=unknown)
  - Bought Out Items
  - Excluded Items
  - Summary
"""
import io
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import structlog
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.bom_builder import resolve_display_name
from app.services.bom_manifest import (
    MANIFEST_HEADERS,
    _build_node_to_ref,
    assign_bom_items,
    build_manifest_rows,
    build_ref_to_stl,
    resolve_classification_ref,
)
from app.services.stl_thumbnail import render_stl_thumbnail

logger = structlog.get_logger()

STEEL_DENSITY = 7.85e-6  # kg/mm³

# Styling constants
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_BODY_FONT = Font(name="Calibri", size=9)
_BODY_FONT_BOLD = Font(name="Calibri", size=9, bold=True)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
_THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# Category fill colours (light tints matching the frontend)
_CAT_FILLS = {
    "cnc":      PatternFill(start_color="DCEAFB", end_color="DCEAFB", fill_type="solid"),  # blue
    "formed":   PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # orange
    "unknown":  PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),  # amber
    "bo":       PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),  # green
    "excluded": PatternFill(start_color="E8D5F5", end_color="E8D5F5", fill_type="solid"),  # purple
}

# Confidence-level cell fills
_CONFIDENCE_FILLS = {
    "Excellent": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # green
    "Good":      PatternFill(start_color="D6EEFA", end_color="D6EEFA", fill_type="solid"),  # light blue
    "Marginal":  PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # yellow/amber
    "Review":    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # red/pink
}
_CONFIDENCE_FONTS = {
    "Excellent": Font(name="Calibri", size=9, bold=True, color="006100"),
    "Good":      Font(name="Calibri", size=9, bold=True, color="1F4E79"),
    "Marginal":  Font(name="Calibri", size=9, bold=True, color="9C6500"),
    "Review":    Font(name="Calibri", size=9, bold=True, color="9C0006"),
}


def _round(v, dp=1):
    if v is None or not isinstance(v, (int, float)):
        return v
    return round(v, dp)


# ---------------------------------------------------------------
# Data assembly — mirrors the frontend _downloadBOM logic
# ---------------------------------------------------------------

def _apply_refined_overrides(cnc_results: dict, overrides: dict) -> None:
    """Overwrite refined_class on CNC results from user overrides (in place).

    Keys are ``ref_id`` (single-solid) or ``ref_id:sN`` (a multi-solid
    sub-solid).  A user override is authoritative, so refined_confidence is set
    to 1.0 to clear the low-confidence review flag.  Mutates the request-local
    cache only — never persisted, never touches the base ``type`` or NC files.
    """
    import re as _re
    for key, code in (overrides or {}).items():
        if not code:
            continue
        m = _re.match(r"^(.*):s(\d+)$", key)
        if m:
            ref, idx = m.group(1), int(m.group(2))
            r = cnc_results.get(ref)
            solids = r.get("solids") if isinstance(r, dict) else None
            if solids and 0 <= idx < len(solids) and isinstance(solids[idx], dict):
                solids[idx]["refined_class"] = code
                solids[idx]["refined_confidence"] = 1.0
        else:
            r = cnc_results.get(key)
            if isinstance(r, dict):
                r["refined_class"] = code
                r["refined_confidence"] = 1.0


def _build_bom_data(cache: dict) -> dict:
    """
    Extract BOM items from the sidecar cache, applying consolidation
    groups so the output matches the visual BOM in the frontend.
    """
    cnc_results: Dict[str, dict] = cache.get("cnc_analysis", {})
    member_names: Dict[str, str] = cache.get("cnc_member_names", {})
    parent_names: Dict[str, str] = cache.get("cnc_parent_names", {})
    classifications: Dict[str, str] = cache.get("project_state", {}).get("classifications", {})
    stl_map_raw: Dict[str, str] = cache.get("project_state", {}).get("stl_map", {}) or {}
    consolidation: List[dict] = cache.get("consolidation", {}).get("part_groups", [])

    # Apply user overrides of the auto-detected refined sub-class onto the
    # (request-local) CNC results so the Refined Class column reflects them.
    # Labelling only — base type/NC files are untouched.
    _apply_refined_overrides(
        cnc_results, cache.get("project_state", {}).get("refined_overrides", {})
    )

    project_number = cache.get("cnc_project_number", "")
    steel_grade = cache.get("cnc_steel_grade", "")

    # Build ref_id → consolidation group lookup
    ref_to_group: Dict[str, dict] = {}
    for g in consolidation:
        for rid in g.get("ref_ids", []):
            ref_to_group[rid] = g

    # Shared BOM Item numbering — same source of truth as manifest.csv so the
    # 'BOM Item' column on this sheet matches the manifest row IDs.
    bom_assignments = assign_bom_items(cache)
    # Resolve instance node IDs (e.g. 0:1:1:3:1) → prototype ref_id so the
    # row iteration below finds cnc_results/member_names entries that are
    # keyed by ref_id.  stl_map is the same shape — invert it once for the
    # thumbnail lookup.
    tree = ((cache.get("analysis") or {}).get("assembly_tree") or [])
    node_to_ref = _build_node_to_ref(tree)
    stl_map = build_ref_to_stl(stl_map_raw, tree)

    # Categorise classified refs
    cnc_items: List[dict] = []
    formed_items: List[dict] = []  # formed plate / formed section — no NC/DXF yet
    unknown_items: List[dict] = []
    bo_items: List[dict] = []
    excluded_items: List[dict] = []

    seen_refs: set = set()  # track refs we've already emitted a row for

    for node_id, action in classifications.items():
        # Classifications are stored against instance node IDs (and "<ref>:s<n>"
        # solid keys for per-solid classified multi-solid parts); resolve to the
        # prototype ref_id so the lookups against cnc_results, member_names,
        # parent_names, and stl_map (all keyed by ref_id) succeed.
        ref_id = resolve_classification_ref(node_id, node_to_ref)
        if ref_id in seen_refs:
            continue
        name = member_names.get(ref_id, ref_id)
        parent = parent_names.get(ref_id, "")
        stl_url = stl_map.get(ref_id, "")
        cnc_res = cnc_results.get(ref_id)

        # Check consolidation group
        group = ref_to_group.get(ref_id)
        if group:
            # Mark all refs in the group as seen so later instance
            # classifications for any group member don't double-count.
            for rid in group["ref_ids"]:
                seen_refs.add(rid)
            # Use canonical name and sum qty across group
            name = group.get("canonical_name", name)
            qty = group.get("total_count", len(group.get("ref_ids", [])))
            # Use first ref's STL/CNC for the thumbnail/data
            first_ref = group["ref_ids"][0]
            stl_url = stl_map.get(first_ref, stl_url)
            cnc_res = cnc_results.get(first_ref, cnc_res)
        else:
            seen_refs.add(ref_id)
            # Singleton ref — qty comes from the BOM assignment (per-ref
            # instance count from the tree), keeping the BOM in lockstep with
            # the manifest and the NC1 quantity field.
            qty = (bom_assignments.get(ref_id) or {}).get("bom_total_qty", 1)

        # Recover the parent assembly name for generic SOLID/COMPOUND leaves so
        # the BOM is recognisable against the tree.  Excluded artifacts keep
        # their generic name to stay distinct from their solid-bearing sibling.
        display = name if action == "exclude" else resolve_display_name(name, parent)

        item = {
            "bom_item": (bom_assignments.get(ref_id) or {}).get("bom_item", ""),
            "name": display,
            "ref_id": ref_id,
            "qty": qty,
            "parent": parent,
            "stl_url": stl_url,
            "cnc": cnc_res,
        }

        if action == "postprocess":
            # Formed plate / formed (bent) section: a flat-pattern development is
            # needed before NC1/DXF can be produced, so keep them in their own
            # bucket rather than mixed with ready-to-cut CNC items. The refined
            # class reflects any user override (applied above).
            refined = (cnc_res or {}).get("refined_class")
            if refined in ("FORMED_PLATE", "BENT_SECTION"):
                formed_items.append(item)
            elif cnc_res and cnc_res.get("type") not in (None, "unknown"):
                cnc_items.append(item)
            else:
                unknown_items.append(item)
        elif action == "bought-out":
            bo_items.append(item)
        elif action == "exclude":
            excluded_items.append(item)

    return {
        "project_number": project_number,
        "steel_grade": steel_grade,
        "cnc_items": cnc_items,
        "formed_items": formed_items,
        "unknown_items": unknown_items,
        "bo_items": bo_items,
        "excluded_items": excluded_items,
    }


# ---------------------------------------------------------------
# Thumbnail cache (per-request in-memory)
# ---------------------------------------------------------------

def _resolve_stl_path(stl_url: str) -> Optional[Path]:
    """Convert a relative STL URL to an absolute filesystem path."""
    if not stl_url:
        return None
    # stl_url is like /outputs/stl/9dd0ba2e/PartName.stl
    # On disk it lives at /app/outputs/stl/...
    if stl_url.startswith("/outputs/"):
        return Path("/app") / stl_url.lstrip("/")
    return None


# ---------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------

def _style_header(ws, col_count: int, row: int = 1):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _ALIGN_CENTER
        cell.border = _THIN_BORDER


def _style_cell(cell, align=_ALIGN_LEFT, bold=False):
    cell.font = _BODY_FONT_BOLD if bold else _BODY_FONT
    cell.alignment = align
    cell.border = _THIN_BORDER


def _write_cnc_sheet(
    wb: Workbook,
    sheet_name: str,
    items: List[dict],
    category_fill: PatternFill,
    thumb_cache: dict,
):
    """Write a sheet for CNC items (sections/plates) with thumbnails."""
    ws = wb.create_sheet(sheet_name)

    headers = [
        "Thumbnail", "BOM Item", "Name", "Qty", "Type", "Category", "Designation",
        "Profile", "L (mm)", "H (mm)", "W (mm)", "T (mm)",
        "Mass Each (kg)", "Total Mass (kg)", "Confidence",
        "Holes", "End Cuts",
        "DXF", "NC1", "Parent Assembly",
        "Refined Class", "Class Conf.", "Review?",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, len(headers))

    # Column widths
    col_widths = [30, 9, 30, 6, 8, 12, 18, 6, 10, 10, 10, 8, 12, 12, 12, 6, 8, 16, 16, 25,
                  16, 10, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thumb_col = 1
    row_height = 80  # pixels for thumbnail rows

    for row_idx, item in enumerate(items, 2):
        cnc = item.get("cnc") or {}
        dims = cnc.get("dims") or {}

        mass_each = cnc.get("mass_kg")
        if mass_each is None and cnc.get("volume_mm3"):
            mass_each = _round(cnc["volume_mm3"] * STEEL_DENSITY, 3)

        dxf_name = Path(cnc["dxf_path"]).name if cnc.get("dxf_path") else ""
        nc1_name = Path(cnc["nc1_path"]).name if cnc.get("nc1_path") else ""

        # For unknowns, show measured CSA in the designation column
        designation = cnc.get("designation", "")
        if cnc.get("type") == "unknown" and not designation:
            csa = cnc.get("section_area")
            msg = cnc.get("message", "")
            parts = []
            if csa:
                parts.append(f"CSA={csa}mm²")
            if msg:
                parts.append(msg)
            designation = " | ".join(parts) if parts else "UNMATCHED"

        confidence = cnc.get("confidence", "")

        # Refined decision-tree class (rw7.1): formed/bent/exclude that the base
        # type can't express, plus a review flag for low-confidence calls.
        refined = cnc.get("refined_class", "")
        refined_conf = cnc.get("refined_confidence")
        review = "REVIEW" if (refined_conf is not None and refined_conf < 0.5) else ""

        values = [
            None,  # thumbnail column — filled separately
            item.get("bom_item", ""),
            item["name"],
            item["qty"],
            cnc.get("type", ""),
            cnc.get("category", ""),
            designation,
            cnc.get("profile_type", ""),
            _round(dims.get("L")),
            _round(dims.get("H")),
            _round(dims.get("W")),
            _round(dims.get("T")),
            _round(mass_each, 3),
            _round(mass_each * item["qty"], 3) if mass_each else None,
            confidence,
            cnc.get("holes", ""),
            "Yes" if cnc.get("end_cuts") else "",
            dxf_name,
            nc1_name,
            item.get("parent", ""),
            refined,
            refined_conf if refined_conf is not None else "",
            review,
        ]

        confidence_col = 15  # 1-based column index for Confidence

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            _style_cell(cell, align=_ALIGN_RIGHT if col >= 4 and col <= 14 else _ALIGN_LEFT)
            if col >= 2:
                cell.fill = category_fill
            # Override confidence cell with colour-coded styling
            if col == confidence_col and confidence:
                cell.fill = _CONFIDENCE_FILLS.get(confidence, category_fill)
                cell.font = _CONFIDENCE_FONTS.get(confidence, _BODY_FONT_BOLD)
                cell.alignment = _ALIGN_CENTER

        # Set row height for thumbnail
        ws.row_dimensions[row_idx].height = 115

        # Insert thumbnail
        stl_path = _resolve_stl_path(item.get("stl_url", ""))
        if stl_path:
            cache_key = str(stl_path)
            if cache_key not in thumb_cache:
                thumb_cache[cache_key] = render_stl_thumbnail(stl_path, width=320, height=240)
            png_bytes = thumb_cache[cache_key]
            if png_bytes:
                img = XlImage(io.BytesIO(png_bytes))
                img.width = 200
                img.height = 150
                cell_ref = f"A{row_idx}"
                ws.add_image(img, cell_ref)

    # Freeze top row
    ws.freeze_panes = "B2"


def _write_simple_sheet(
    wb: Workbook,
    sheet_name: str,
    items: List[dict],
    category_fill: PatternFill,
    thumb_cache: dict,
):
    """Write a sheet for bought-out or excluded items (minimal columns)."""
    ws = wb.create_sheet(sheet_name)

    headers = ["Thumbnail", "BOM Item", "Name", "Qty", "Parent Assembly"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, len(headers))

    col_widths = [30, 9, 35, 8, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row_idx, item in enumerate(items, 2):
        values = [None, item.get("bom_item", ""), item["name"], item["qty"], item.get("parent", "")]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            _style_cell(cell, align=_ALIGN_RIGHT if col == 4 else _ALIGN_LEFT)
            if col >= 2:
                cell.fill = category_fill

        ws.row_dimensions[row_idx].height = 115

        stl_path = _resolve_stl_path(item.get("stl_url", ""))
        if stl_path:
            cache_key = str(stl_path)
            if cache_key not in thumb_cache:
                thumb_cache[cache_key] = render_stl_thumbnail(stl_path, width=320, height=240)
            png_bytes = thumb_cache[cache_key]
            if png_bytes:
                img = XlImage(io.BytesIO(png_bytes))
                img.width = 200
                img.height = 150
                ws.add_image(img, f"A{row_idx}")

    ws.freeze_panes = "B2"


def _write_nc_index_sheet(wb: Workbook, manifest_rows: List[dict]) -> None:
    """One row per generated NC1/DXF file linking back to BOM Item #."""
    ws = wb.create_sheet("NC Index")

    headers = [
        "BOM Item", "Filename", "Ext", "Qty",
        "BOM Total Qty", "Hash",
        "Consolidation Group", "Member Name", "Parent Assembly",
        "Type", "Category", "Designation", "Profile",
        "L (mm)", "H (mm)", "W (mm)", "T (mm)",
        "Mass Each (kg)", "Ref ID", "Solid Idx",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, len(headers))

    col_widths = [9, 38, 5, 6, 12, 32, 24, 28, 25, 10, 9, 18, 8, 9, 9, 9, 8, 12, 12, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Default fill colour matches CNC sheet so it reads as part of the BOM
    fill = _CAT_FILLS["cnc"]

    for row_idx, r in enumerate(manifest_rows, 2):
        values = [
            r.get("bom_item", ""),
            r.get("filename", ""),
            r.get("ext", ""),
            r.get("qty", ""),
            r.get("bom_total_qty", ""),
            r.get("nc1_hash", ""),
            r.get("consolidation_group", ""),
            r.get("member_name", ""),
            r.get("parent_assembly", ""),
            r.get("type", ""),
            r.get("category", ""),
            r.get("designation", ""),
            r.get("profile_type", ""),
            _round(r.get("L_mm")),
            _round(r.get("H_mm")),
            _round(r.get("W_mm")),
            _round(r.get("T_mm")),
            _round(r.get("mass_kg"), 3),
            r.get("ref_id", ""),
            r.get("solid_idx", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            _style_cell(cell, align=_ALIGN_RIGHT if col in (4, 5, 14, 15, 16, 17, 18, 20) else _ALIGN_LEFT)
            cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}{max(2, len(manifest_rows) + 1)}"
    )


def _compute_assembly_bbox(stl_dir: Path) -> Optional[Tuple[float, float, float]]:
    """Compute overall bounding box (X, Y, Z) in mm from all STLs in the directory."""
    if not stl_dir.is_dir():
        return None
    stl_files = sorted(stl_dir.glob("*.stl"))
    if not stl_files:
        return None
    try:
        import numpy as np
        import trimesh
        all_min = np.array([np.inf, np.inf, np.inf])
        all_max = np.array([-np.inf, -np.inf, -np.inf])
        loaded = False
        for f in stl_files:
            try:
                mesh = trimesh.load_mesh(str(f))
                if mesh.is_empty:
                    continue
                all_min = np.minimum(all_min, mesh.vertices.min(axis=0))
                all_max = np.maximum(all_max, mesh.vertices.max(axis=0))
                loaded = True
            except Exception:
                continue
        if not loaded:
            return None
        extents = all_max - all_min
        return (round(float(extents[0]), 1),
                round(float(extents[1]), 1),
                round(float(extents[2]), 1))
    except Exception:
        return None


def _write_summary_sheet(wb: Workbook, bom: dict, filename: str,
                         assembly_thumb_path: Optional[Path] = None,
                         stl_dir: Optional[Path] = None):
    """Write a summary sheet with project info, totals, and assembly thumbnail."""
    ws = wb.create_sheet("Summary")

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40

    # Compute overall bounding box from STL meshes
    bbox = _compute_assembly_bbox(stl_dir) if stl_dir else None

    rows = [
        ("Project Number", bom.get("project_number", "")),
        ("Steel Grade", bom.get("steel_grade", "")),
        ("STEP File", filename),
        ("Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("", ""),
    ]

    if bbox:
        sorted_dims = sorted(bbox, reverse=True)
        rows.append(("Overall Dimensions (mm)",
                      f"{sorted_dims[0]} x {sorted_dims[1]} x {sorted_dims[2]}"))
        rows.append(("", ""))

    rows.extend([
        ("CNC Items (types)", len(bom["cnc_items"])),
        ("CNC Items (total qty)", sum(i["qty"] for i in bom["cnc_items"])),
        ("CNC Total Weight (kg)", _round(sum(
            (i.get("cnc") or {}).get("mass_kg", 0) * i["qty"]
            for i in bom["cnc_items"]
            if (i.get("cnc") or {}).get("mass_kg")
        ), 1)),
        ("", ""),
        ("Formed Plate (types)", len(bom.get("formed_items", []))),
        ("Formed Plate (total qty)", sum(i["qty"] for i in bom.get("formed_items", []))),
        ("", ""),
        ("Unknown Items (types)", len(bom["unknown_items"])),
        ("Unknown Items (total qty)", sum(i["qty"] for i in bom["unknown_items"])),
        ("", ""),
        ("Bought Out Items (types)", len(bom["bo_items"])),
        ("Bought Out Items (total qty)", sum(i["qty"] for i in bom["bo_items"])),
        ("", ""),
        ("Excluded Items (types)", len(bom["excluded_items"])),
        ("Excluded Items (total qty)", sum(i["qty"] for i in bom["excluded_items"])),
    ])

    for row_idx, (label, value) in enumerate(rows, 1):
        cell_a = ws.cell(row=row_idx, column=1, value=label)
        cell_b = ws.cell(row=row_idx, column=2, value=value)
        cell_a.font = _BODY_FONT_BOLD
        cell_b.font = _BODY_FONT
        cell_a.alignment = _ALIGN_LEFT
        cell_b.alignment = _ALIGN_LEFT

    # Embed assembly thumbnail to the right of the text, starting at column C
    if assembly_thumb_path and assembly_thumb_path.exists():
        try:
            img = XlImage(str(assembly_thumb_path))
            # Read native pixel dimensions and scale to max 640px wide,
            # preserving aspect ratio
            native_w = img.width
            native_h = img.height
            max_w = 640
            if native_w > 0 and native_h > 0:
                scale = min(max_w / native_w, 1.0)
                img.width = int(native_w * scale)
                img.height = int(native_h * scale)
            else:
                img.width = max_w
                img.height = 480
            # Place at C1 (to the right of A/B summary columns)
            ws.add_image(img, "C1")
            # Widen column C so the image isn't clipped
            ws.column_dimensions["C"].width = 90
        except Exception:
            pass  # skip thumbnail if anything goes wrong


# ---------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------

def generate_bom_xlsx(filename: str) -> Optional[bytes]:
    """
    Generate an Excel BOM workbook for the given STEP filename.

    Returns the .xlsx file as bytes, or None if no data is available.
    """
    from app.routers.cnc_analysis import _load_cache

    cache = _load_cache(filename)
    if not cache or "cnc_analysis" not in cache:
        return None

    bom = _build_bom_data(cache)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    thumb_cache: dict = {}  # stl_path → png bytes (or None)

    if bom["cnc_items"]:
        _write_cnc_sheet(wb, "CNC Items", bom["cnc_items"],
                         _CAT_FILLS["cnc"], thumb_cache)

    if bom.get("formed_items"):
        # Formed plate / formed section — no NC1/DXF until flat-pattern
        # development exists, so they get their own sheet (full detail +
        # thumbnails; the DXF/NC1 columns are intentionally blank).
        _write_cnc_sheet(wb, "Formed Plate", bom["formed_items"],
                         _CAT_FILLS["formed"], thumb_cache)

    if bom["unknown_items"]:
        _write_cnc_sheet(wb, "Unknown Items", bom["unknown_items"],
                         _CAT_FILLS["unknown"], thumb_cache)

    if bom["bo_items"]:
        _write_simple_sheet(wb, "Bought Out", bom["bo_items"],
                            _CAT_FILLS["bo"], thumb_cache)

    if bom["excluded_items"]:
        _write_simple_sheet(wb, "Excluded", bom["excluded_items"],
                            _CAT_FILLS["excluded"], thumb_cache)

    # NC Index — one row per generated NC1/DXF file with BOM linkage
    manifest_rows = build_manifest_rows(cache)
    if manifest_rows:
        _write_nc_index_sheet(wb, manifest_rows)

    # Resolve assembly thumbnail and STL directory for bounding box
    run_id = filename[:8]
    stl_dir = Path("/app/outputs/stl") / run_id
    assembly_thumb = stl_dir / "_thumbnail.png"
    _write_summary_sheet(wb, bom, filename,
                         assembly_thumb_path=assembly_thumb,
                         stl_dir=stl_dir)

    # Ensure at least one visible sheet
    if len(wb.sheetnames) == 0:
        return None

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
