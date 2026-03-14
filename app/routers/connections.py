"""
Connection detection router.

Detects weld and bolt connections between solids in a STEP assembly.

Endpoints
---------
POST /connections/detect/{filename}
     Body: {"node_id": "0:1:1:9"}   (optional — scope to a subtree)
     Start a background worker subprocess.
     Returns {"task_id": "...", "status": "pending"}.

GET  /connections/status/{task_id}
     Poll a running detection task.

GET  /connections/result/{filename}
     Return cached connection results from sidecar JSON.

POST /connections/detect-all/{filename}
     Batch-detect connections across all assembly levels.

GET  /connections/detect-all-status/{task_id}
     Poll batch detection progress.

GET  /connections/tree-status/{filename}
     Return per-node detection/verification status for tree badges.

POST /connections/preview/{filename}
     Body: {"node_id": "0:1:1:8:1"}
     Generate per-solid STLs in world coordinates for 3D preview.

GET  /connections/preview-status/{task_id}
     Poll preview STL generation task.  Returns STL file list when done.
"""
import asyncio
import csv
import hashlib
import io
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Body, HTTPException, status
from fastapi.responses import StreamingResponse
import structlog

from app.config import settings
from app.services.task_manager import task_manager, TaskStatus

_WORKER = Path(__file__).parent.parent / "workers" / "detect_connections.py"
_STL_WORKER = Path(__file__).parent.parent / "workers" / "generate_stl.py"
_TIMEOUT = 1800  # 30 minutes
_STL_TIMEOUT = 600  # 10 minutes

router = APIRouter()
logger = structlog.get_logger()


# ── Path / cache helpers ──────────────────────────────────────────────────

def _analysis_json_path(filename: str) -> Path:
    stem = Path(filename).stem
    return Path(settings.ANALYSIS_OUTPUT_DIR) / f"{stem}.json"


def _load_cache(filename: str) -> Optional[dict]:
    path = _analysis_json_path(filename)
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            pass
    return None


def _save_connections(
    filename: str, data: dict, node_id: str = "", scope: str = "all",
) -> None:
    """Merge connection results into the sidecar JSON.

    Results are stored per ``node_id|scope`` key so that switching between
    assembly nodes does not overwrite previous detection runs.
    """
    path = _analysis_json_path(filename)
    path.parent.mkdir(parents=True, exist_ok=True)

    existing: dict = {}
    if path.exists():
        try:
            existing = json.loads(path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            pass

    entry = {
        "detected_at": datetime.now(timezone.utc).isoformat(),
        "node_id": node_id,
        "scope": scope,
        **data,
    }

    # Per-key store so previous runs are preserved
    store = existing.setdefault("connections_store", {})
    key = f"{node_id}|{scope}"
    store[key] = entry

    # Also keep latest in "connections" for backward compat
    existing["connections"] = entry
    path.write_text(json.dumps(existing, indent=2), encoding="utf-8")
    logger.info("connections_saved", filename=filename, key=key)


# ── Endpoints ─────────────────────────────────────────────────────────────

@router.post("/connections/detect/{filename}")
async def detect_connections(
    filename: str,
    body: Dict[str, Any] = Body(default={}),
) -> Dict[str, Any]:
    """Launch connection detection as a background subprocess."""
    file_path = Path(settings.UPLOAD_DIR) / filename
    if not file_path.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error": "File not found", "filename": filename},
        )

    analysis_json = _analysis_json_path(filename)
    if not analysis_json.exists():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "error": "Assembly tree not found — run analysis first",
                "filename": filename,
            },
        )

    node_id: Optional[str] = body.get("node_id")
    scope: str = body.get("scope", "all")
    if scope not in ("all", "siblings", "within-part", "cross-assembly"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error": f"Invalid scope: {scope!r}"},
        )

    # Check for existing in-progress task
    existing_tasks = task_manager.get_tasks_for_file(
        filename, task_type="connection_detection",
    )
    for task in existing_tasks:
        if task.status in (TaskStatus.PENDING, TaskStatus.RUNNING):
            logger.info(
                "connections_task_already_running",
                filename=filename,
                task_id=task.task_id,
            )
            return {"task_id": task.task_id, "status": task.status.value}

    # Build worker argv — always pass node_id slot (empty string = no scope)
    worker_args = [
        sys.executable,
        str(_WORKER),
        str(file_path),
        str(analysis_json),
        node_id or "",
        scope,
    ]

    async def run_async(_progress_callback):
        try:
            proc = await asyncio.create_subprocess_exec(
                *worker_args,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
        except Exception as exc:
            raise RuntimeError(f"Failed to launch connection worker: {exc}")

        try:
            stdout_bytes, stderr_bytes = await asyncio.wait_for(
                proc.communicate(), timeout=float(_TIMEOUT),
            )
        except asyncio.TimeoutError:
            proc.kill()
            await proc.wait()
            raise RuntimeError(
                f"Connection detection timed out after {_TIMEOUT // 60} min"
            )

        stdout = stdout_bytes.decode("utf-8", errors="replace")
        stderr = stderr_bytes.decode("utf-8", errors="replace").strip()

        if proc.returncode != 0:
            tail = stdout.strip()[-500:] if stdout.strip() else ""
            err = (
                f"Connection worker failed (exit {proc.returncode}): {stderr}"
                if stderr
                else f"Connection worker crashed (exit {proc.returncode}): {tail}"
            )
            logger.error(
                "connections_worker_failed",
                filename=filename,
                returncode=proc.returncode,
                stderr=stderr[:1000] if stderr else "",
            )
            raise RuntimeError(err)

        json_line = next(
            (l for l in stdout.splitlines() if l.strip().startswith("{")),
            stdout,
        )
        payload = json.loads(json_line)
        _save_connections(filename, payload, node_id=node_id or "", scope=scope)
        return payload

    task_id = task_manager.submit_async(
        "connection_detection", filename, run_async,
    )
    logger.info(
        "connections_task_submitted",
        filename=filename,
        task_id=task_id,
        node_id=node_id,
    )
    return {"task_id": task_id, "status": "pending"}


@router.get("/connections/status/{task_id}")
async def get_connection_status(task_id: str) -> Dict[str, Any]:
    """Poll a connection detection task."""
    info = task_manager.get_task(task_id)
    if not info:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error": "Task not found", "task_id": task_id},
        )

    if info.status == TaskStatus.FAILED:
        return {"status": "failed", "error": info.error or "Detection failed"}

    if info.status in (TaskStatus.PENDING, TaskStatus.RUNNING):
        return {"status": info.status.value, "task_id": task_id}

    results = info.results or {}
    _inject_ids_and_verifications(results)
    return {
        "status": "completed",
        "results": results,
        "filename": info.filename,
    }


@router.get("/connections/result/{filename}")
async def get_connection_result(
    filename: str,
    node_id: Optional[str] = None,
    scope: Optional[str] = None,
) -> Dict[str, Any]:
    """Return cached connection results with IDs and verification state.

    If *node_id* and *scope* are given, look up results for that specific
    detection run from ``connections_store``.  Otherwise return the most
    recent run stored under ``connections``.
    """
    cache = _load_cache(filename)
    if not cache:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={
                "error": "No connection results found — run detection first",
                "filename": filename,
            },
        )

    data = None
    if node_id is not None:
        store = cache.get("connections_store", {})
        if scope is not None:
            # Try exact match first
            data = store.get(f"{node_id}|{scope}")
        if data is None:
            # Fall back: find any scope for this node_id
            # Priority: siblings > within-part > all
            _scope_priority = {"siblings": 3, "within-part": 2, "all": 1}
            best = None
            best_pri = -1
            for key, entry in store.items():
                if entry.get("node_id") == node_id:
                    pri = _scope_priority.get(entry.get("scope", ""), 0)
                    if pri > best_pri:
                        best = entry
                        best_pri = pri
            data = best
    else:
        # No specific params — return latest run
        data = cache.get("connections")
    if data is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={
                "error": "No connection results found — run detection first",
                "filename": filename,
            },
        )
    return _inject_ids_and_verifications(data)


# ── Connection IDs & Verification ──────────────────────────────────────────

def _connection_id(conn: dict) -> str:
    """Stable identifier for a connection (hash of the sorted solid pair)."""
    a = conn.get("solid_a", {})
    b = conn.get("solid_b", {})
    key_a = f"{a.get('node_id', '')}:{a.get('solid_index', 0)}"
    key_b = f"{b.get('node_id', '')}:{b.get('solid_index', 0)}"
    pair = ":".join(sorted([key_a, key_b]))
    return hashlib.sha256(pair.encode()).hexdigest()[:12]


def _inject_ids_and_verifications(data: dict) -> dict:
    """Add ``id`` to each connection and merge stored verifications."""
    verifications = data.get("verifications", {})
    for conn in data.get("connections", []):
        cid = _connection_id(conn)
        conn["id"] = cid
        v = verifications.get(cid)
        if v:
            conn["verification"] = v
    return data


@router.put("/connections/verify/{filename}")
async def verify_connection(
    filename: str,
    body: Dict[str, Any] = Body(...),
) -> Dict[str, Any]:
    """Accept, reject, or reclassify a single connection.

    Body::

        {
            "connection_id": "abc123def456",
            "action": "accept" | "reject" | "reclassify",
            "new_type": "welded" | "bolted" | "contact"   # only for reclassify
        }
    """
    connection_id: str = body.get("connection_id", "")
    action: str = body.get("action", "")
    new_type: Optional[str] = body.get("new_type")

    if not connection_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error": "connection_id is required"},
        )
    if action not in ("accept", "reject", "reclassify"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error": f"Invalid action: {action!r}"},
        )
    if action == "reclassify" and new_type not in ("welded", "bolted", "contact"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error": f"Invalid new_type for reclassify: {new_type!r}"},
        )

    path = _analysis_json_path(filename)
    if not path.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error": "No analysis found", "filename": filename},
        )

    existing = json.loads(path.read_text(encoding="utf-8"))
    connections_section = existing.get("connections", {})
    verifications: dict = connections_section.setdefault("verifications", {})

    verification_entry: Dict[str, Any] = {
        "action": action,
        "at": datetime.now(timezone.utc).isoformat(),
    }
    if action == "reclassify":
        verification_entry["new_type"] = new_type

    verifications[connection_id] = verification_entry
    path.write_text(json.dumps(existing, indent=2), encoding="utf-8")

    logger.info(
        "connection_verified",
        filename=filename,
        connection_id=connection_id,
        action=action,
    )
    return {"status": "ok", "connection_id": connection_id, "action": action}


# ── Export ────────────────────────────────────────────────────────────────

@router.get("/connections/export/{filename}")
async def export_connections_csv(filename: str):
    """Export connection results as a CSV file."""
    cache = _load_cache(filename)
    if not cache or "connections" not in cache:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error": "No connection results — run detection first"},
        )

    data = _inject_ids_and_verifications(cache["connections"])
    connections = data.get("connections", [])

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "#", "Part A", "Part A Node", "Solid A Index",
        "Part B", "Part B Node", "Solid B Index",
        "Type", "Weld Length (mm)", "Bolt Count", "Bolt Diameter (mm)",
        "Min Distance (mm)", "Contact Subtype", "Auto Classified",
        "Verification", "Reclassified Type",
    ])

    for i, c in enumerate(connections):
        a = c.get("solid_a", {})
        b = c.get("solid_b", {})
        v = c.get("verification", {})

        # Effective type (after reclassification)
        eff_type = c.get("type", "")
        if v.get("action") == "reclassify" and v.get("new_type"):
            eff_type = v["new_type"]

        writer.writerow([
            i + 1,
            a.get("name", ""),
            a.get("node_id", ""),
            a.get("solid_index", ""),
            b.get("name", ""),
            b.get("node_id", ""),
            b.get("solid_index", ""),
            eff_type,
            c.get("weld_length_mm", ""),
            c.get("bolt_count", ""),
            c.get("bolt_diameter_mm", ""),
            c.get("min_distance_mm", ""),
            c.get("contact_subtype", ""),
            "Yes" if c.get("auto_classified") else "",
            v.get("action", ""),
            v.get("new_type", ""),
        ])

    stem = Path(filename).stem
    csv_filename = f"{stem}_connections.csv"

    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{csv_filename}"'},
    )


@router.get("/connections/export-xlsx/{filename}")
async def export_connections_xlsx(
    filename: str,
    node_id: Optional[str] = None,
    scope: Optional[str] = None,
) -> StreamingResponse:
    """Export connection results as an Excel (.xlsx) file."""
    cache = _load_cache(filename)
    if not cache:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error": "No connection results — run detection first"},
        )

    # Resolve the right result set (same logic as get_connection_result)
    data = None
    if node_id is not None:
        store = cache.get("connections_store", {})
        key = f"{node_id}|{scope or 'all'}"
        data = store.get(key)
        if data is None:
            for entry in store.values():
                if entry.get("node_id") == node_id:
                    data = entry
                    break
    if data is None:
        data = cache.get("connections")
    if not data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error": "No connection results — run detection first"},
        )

    _inject_ids_and_verifications(data)
    connections = data.get("connections", [])
    summary = data.get("summary", {})
    detected_at = data.get("detected_at", "")

    # ── Build workbook ────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Connections"

    # Styles
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    verified_fill = PatternFill("solid", fgColor="D6F4E0")   # green tint
    reclass_fill = PatternFill("solid", fgColor="DBEAFE")    # blue tint
    centre = Alignment(horizontal="center")

    # ── Summary block (rows 1–3) ──────────────────────────────────────────
    stem = Path(filename).stem
    ws.append([f"Connections — {stem}"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([
        f"Detected: {detected_at[:19].replace('T', ' ') if detected_at else '—'}",
        "",
        f"Welded: {summary.get('welded_connections', 0)}",
        f"Bolted: {summary.get('bolted_connections', 0)}",
        f"Contacts: {summary.get('unclassified_contacts', 0)}",
        f"Weld length: {summary.get('total_weld_length_mm', 0):.1f} mm",
    ])
    ws.append([])  # blank row

    # ── Header row ────────────────────────────────────────────────────────
    headers = [
        "#", "Part A", "Part B", "Type",
        "Weld Length (mm)", "Bolt Count", "Bolt Ø (mm)",
        "Min Gap (mm)", "Contact Subtype", "Auto", "Verified", "Notes",
    ]
    ws.append(headers)
    hdr_row = ws.max_row
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(hdr_row, col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = centre

    # ── Data rows ─────────────────────────────────────────────────────────
    for i, c in enumerate(connections, 1):
        a = c.get("solid_a", {})
        b = c.get("solid_b", {})
        v = c.get("verification", {}) or {}

        eff_type = c.get("type", "")
        if v.get("action") == "reclassify" and v.get("new_type"):
            eff_type = v["new_type"]

        name_a = a.get("name", "")
        if a.get("solid_index", 0):
            name_a += f" [s{a['solid_index']}]"
        name_b = b.get("name", "")
        if b.get("solid_index", 0):
            name_b += f" [s{b['solid_index']}]"

        notes = ""
        if v.get("action") == "reclassify" and v.get("new_type") != c.get("type"):
            notes = f"Reclassified from {c.get('type', '')}"
        elif v.get("action") == "accept":
            notes = "Accepted"

        row = [
            i,
            name_a,
            name_b,
            eff_type.capitalize() if eff_type else "",
            c.get("weld_length_mm") or "",
            c.get("bolt_count") or "",
            c.get("bolt_diameter_mm") or "",
            c.get("min_distance_mm") if c.get("min_distance_mm") is not None else "",
            c.get("contact_subtype") or "",
            "Yes" if c.get("auto_classified") else "",
            v.get("action", "").capitalize() if v.get("action") else "",
            notes,
        ]
        ws.append(row)
        data_row = ws.max_row

        # Row background based on verification state
        if v.get("action") == "reclassify":
            fill = reclass_fill
        elif v.get("action"):
            fill = verified_fill
        else:
            fill = None

        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(data_row, col).fill = fill

    # ── Column widths ─────────────────────────────────────────────────────
    col_widths = [5, 32, 32, 12, 18, 12, 12, 14, 18, 8, 12, 28]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = f"A{hdr_row + 1}"

    # ── Stream response ───────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    xlsx_filename = f"{stem}_connections.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{xlsx_filename}"'},
    )


# ── Batch detection ───────────────────────────────────────────────────────

_MAX_SOLIDS_PER_UNIT = 1000  # mirrors connection_detector.MAX_SOLIDS


def _plan_work_units(
    tree_nodes: list,
    existing_keys: set,
) -> List[Dict[str, Any]]:
    """Walk the assembly tree and plan detection work units.

    Returns a list of ``{"node_id", "scope", "name", "solid_count"}`` dicts.
    Skips units whose ``node_id|scope`` key is already in *existing_keys*.

    Unit priority order:
    1. Cross-assembly (root-level pairwise boundary detection)
    2. Siblings (shallowest depth first)
    3. Within-part
    """
    cross_units: List[Dict[str, Any]] = []
    siblings_units: List[Dict[str, Any]] = []
    within_units: List[Dict[str, Any]] = []

    def _descendant_solids(node: dict) -> int:
        """Recursively count all leaf-level solids under *node*."""
        children = node.get("children", [])
        if not children:
            return node.get("solid_count", 0)
        return sum(_descendant_solids(c) for c in children)

    def _walk(nodes: list, depth: int = 0) -> None:
        for node in nodes:
            nid = node.get("id", "")
            name = node.get("name", "")
            children = node.get("children", [])

            if children:
                # Assembly node — consider siblings detection.
                # Compute total descendant solids (not just direct children).
                total_solids = _descendant_solids(node)
                if 0 < total_solids <= _MAX_SOLIDS_PER_UNIT:
                    key = f"{nid}|siblings"
                    if key not in existing_keys:
                        siblings_units.append({
                            "node_id": nid,
                            "scope": "siblings",
                            "name": name,
                            "solid_count": total_solids,
                            "depth": depth,
                        })
                elif total_solids > _MAX_SOLIDS_PER_UNIT:
                    # Too large for regular siblings — use cross-assembly
                    key = f"{nid}|cross-assembly"
                    if key not in existing_keys:
                        cross_units.append({
                            "node_id": nid,
                            "scope": "cross-assembly",
                            "name": name,
                            "solid_count": total_solids,
                            "depth": depth,
                        })
                # Recurse into children
                _walk(children, depth + 1)
            else:
                # Leaf node — consider within-part detection
                sc = node.get("solid_count", 0)
                if sc > 1:
                    key = f"{nid}|within-part"
                    if key not in existing_keys:
                        within_units.append({
                            "node_id": nid,
                            "scope": "within-part",
                            "name": name,
                            "solid_count": sc,
                        })

    _walk(tree_nodes)

    # Cross-assembly first, then siblings (shallowest first), then within-part
    siblings_units.sort(key=lambda u: u.get("depth", 0))
    return cross_units + siblings_units + within_units


@router.post("/connections/detect-all/{filename}")
async def batch_detect_connections(
    filename: str,
    body: Dict[str, Any] = Body(default={}),
) -> Dict[str, Any]:
    """Launch batch connection detection across all assembly levels."""
    file_path = Path(settings.UPLOAD_DIR) / filename
    if not file_path.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error": "File not found", "filename": filename},
        )

    analysis_json = _analysis_json_path(filename)
    if not analysis_json.exists():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error": "Assembly tree not found — run analysis first"},
        )

    # Check for existing batch task
    existing_tasks = task_manager.get_tasks_for_file(
        filename, task_type="batch_detection",
    )
    for task in existing_tasks:
        if task.status in (TaskStatus.PENDING, TaskStatus.RUNNING):
            return {"task_id": task.task_id, "status": task.status.value}

    # Load analysis tree and plan units
    cache = _load_cache(filename)
    if not cache:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error": "Could not read analysis data"},
        )
    tree = (cache.get("analysis") or {}).get("assembly_tree", [])
    existing_keys = set((cache.get("connections_store") or {}).keys())

    force = body.get("force", False)
    if force:
        existing_keys = set()

    # Optional: scope to a subtree
    root_node_id: Optional[str] = body.get("node_id")
    target_tree = tree
    if root_node_id:
        target_node = _find_tree_node(tree, root_node_id)
        if target_node:
            target_tree = [target_node]
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"error": f"Node {root_node_id!r} not found in tree"},
            )

    units = _plan_work_units(target_tree, existing_keys)

    # Cap total units to keep runtime practical
    max_units = min(int(body.get("max_units", 200)), 2000)
    total_available = len(units)
    if len(units) > max_units:
        units = units[:max_units]

    if not units:
        return {
            "status": "completed",
            "planned_units": 0,
            "total_available": total_available,
            "message": "All connection detection is up to date",
        }

    max_concurrent = min(int(body.get("max_concurrent", 2)), 4)

    async def run_async(progress_callback):
        sem = asyncio.Semaphore(max_concurrent)
        completed: list = []
        failed: list = []

        for i, unit in enumerate(units):
            if progress_callback:
                progress_callback(
                    len(completed),
                    len(units),
                    f"{unit['name']} ({unit['scope']})",
                )

            async with sem:
                worker_args = [
                    sys.executable,
                    str(_WORKER),
                    str(file_path),
                    str(analysis_json),
                    unit["node_id"],
                    unit["scope"],
                ]
                try:
                    proc = await asyncio.create_subprocess_exec(
                        *worker_args,
                        stdout=asyncio.subprocess.PIPE,
                        stderr=asyncio.subprocess.PIPE,
                    )
                    stdout_bytes, stderr_bytes = await asyncio.wait_for(
                        proc.communicate(), timeout=float(_TIMEOUT),
                    )

                    if proc.returncode != 0:
                        stderr_text = stderr_bytes.decode(
                            "utf-8", errors="replace",
                        ).strip()
                        raise RuntimeError(
                            f"Worker exit {proc.returncode}: {stderr_text[:500]}"
                        )

                    stdout = stdout_bytes.decode("utf-8", errors="replace")
                    json_line = next(
                        (l for l in stdout.splitlines()
                         if l.strip().startswith("{")),
                        stdout,
                    )
                    payload = json.loads(json_line)
                    _save_connections(
                        filename, payload,
                        node_id=unit["node_id"], scope=unit["scope"],
                    )

                    summary = payload.get("summary", {})
                    completed.append({
                        "node_id": unit["node_id"],
                        "scope": unit["scope"],
                        "name": unit["name"],
                        "summary": summary,
                    })
                    logger.info(
                        "batch_unit_completed",
                        filename=filename,
                        node_id=unit["node_id"],
                        scope=unit["scope"],
                        i=i + 1,
                        total=len(units),
                    )

                except Exception as exc:
                    failed.append({
                        "node_id": unit["node_id"],
                        "scope": unit["scope"],
                        "name": unit["name"],
                        "error": str(exc),
                    })
                    logger.error(
                        "batch_unit_failed",
                        filename=filename,
                        node_id=unit["node_id"],
                        scope=unit["scope"],
                        error=str(exc),
                    )

        return {
            "completed_units": completed,
            "failed_units": failed,
            "total": len(units),
        }

    task_id = task_manager.submit_async(
        "batch_detection", filename, run_async,
    )
    logger.info(
        "batch_detection_submitted",
        filename=filename,
        task_id=task_id,
        planned_units=len(units),
    )
    return {
        "task_id": task_id,
        "status": "pending",
        "planned_units": len(units),
        "total_available": total_available,
        "units": [
            {"node_id": u["node_id"], "scope": u["scope"], "name": u["name"]}
            for u in units[:50]  # cap response size
        ],
    }


@router.get("/connections/detect-all-status/{task_id}")
async def get_batch_detection_status(task_id: str) -> Dict[str, Any]:
    """Poll a batch detection task."""
    info = task_manager.get_task(task_id)
    if not info:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error": "Task not found", "task_id": task_id},
        )

    if info.status == TaskStatus.FAILED:
        return {"status": "failed", "error": info.error or "Batch detection failed"}

    if info.status in (TaskStatus.PENDING, TaskStatus.RUNNING):
        return {
            "status": info.status.value,
            "task_id": task_id,
            "progress": {
                "completed": info.completed_items,
                "total": info.total_items,
                "current_unit": info.current_item,
                "percent": info.progress,
            },
        }

    results = info.results or {}
    return {
        "status": "completed",
        "completed_units": results.get("completed_units", []),
        "failed_units": results.get("failed_units", []),
        "total": results.get("total", 0),
    }


# ── Tree status ──────────────────────────────────────────────────────────

@router.get("/connections/tree-status/{filename}")
async def get_tree_status(filename: str) -> Dict[str, Any]:
    """Return per-node connection detection and verification status.

    Used by the frontend to render status badges on the assembly tree.
    """
    cache = _load_cache(filename)
    if not cache:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error": "No analysis found", "filename": filename},
        )

    store = cache.get("connections_store", {})
    # Also check legacy "connections" key
    legacy = cache.get("connections")
    if legacy and not store:
        nid = legacy.get("node_id", "")
        sc = legacy.get("scope", "all")
        if nid:
            store = {f"{nid}|{sc}": legacy}

    node_status: Dict[str, Dict[str, Any]] = {}

    for key, data in store.items():
        nid = data.get("node_id", "")
        if not nid:
            # Try to extract from key
            parts = key.split("|", 1)
            nid = parts[0] if parts else ""
        if not nid:
            continue

        connections = data.get("connections", [])
        verifications = data.get("verifications", {})
        total = len(connections)
        accepted = 0
        rejected = 0
        for conn in connections:
            cid = _connection_id(conn)
            v = verifications.get(cid, {})
            action = v.get("action", "")
            if action == "accept":
                accepted += 1
            elif action == "reject":
                rejected += 1
            elif action == "reclassify":
                accepted += 1  # reclassify counts as verified

        pending = total - accepted - rejected

        if total == 0:
            verification = "none"
        elif pending == 0:
            verification = "complete"
        elif accepted + rejected > 0:
            verification = "partial"
        else:
            verification = "unverified"

        scope = data.get("scope", key.split("|")[-1] if "|" in key else "all")
        status_key = f"{nid}|{scope}"

        node_status[status_key] = {
            "node_id": nid,
            "scope": scope,
            "detection": "detected",
            "verification": verification,
            "counts": {
                "total": total,
                "accepted": accepted,
                "rejected": rejected,
                "pending": pending,
            },
        }

    return {"node_status": node_status}


# ── Preview STL generation ─────────────────────────────────────────────────

def _stl_output_dir(filename: str) -> Path:
    """Build the per-file STL output directory using the 8-char hex run ID."""
    run_id = filename[:8]
    return Path(settings.STL_OUTPUT_DIR) / run_id


_SOLID_RE = re.compile(r"^(.+) - Solid (\d+)\.stl$")


def _find_tree_node(nodes: list, target_id: str) -> Optional[dict]:
    """Recursively find a node in the assembly tree by its ``id``."""
    for node in nodes:
        if node.get("id") == target_id:
            return node
        children = node.get("children")
        if children:
            found = _find_tree_node(children, target_id)
            if found:
                return found
    return None


def _find_existing_solid_stls(
    filename: str, node_id: str,
) -> Optional[List[Dict[str, Any]]]:
    """Return cached per-solid STL metadata if files already exist on disk.

    Looks up the node in the analysis JSON to get ``instance_ref`` (the
    NAUO name used as the STL filename base) and ``solid_count``.  If all
    expected ``{instance_ref} - Solid {N}.stl`` files are present, builds
    a result list with instance-based ``node_id`` values so the frontend
    can map them to connection results.
    """
    cache = _load_cache(filename)
    if not cache:
        return None
    tree = (cache.get("analysis") or {}).get("assembly_tree", [])
    if not tree:
        return None

    tree_node = _find_tree_node(tree, node_id)
    if not tree_node:
        return None

    instance_ref = tree_node.get("instance_ref", "")
    solid_count = tree_node.get("solid_count", 0)
    if not instance_ref or solid_count < 1:
        return None

    output_dir = _stl_output_dir(filename)
    if not output_dir.is_dir():
        return None

    run_id = filename[:8]
    # Look in connpreview/ subdirectory for connection-preview STLs
    world_dir = output_dir / "connpreview"
    results: List[Dict[str, Any]] = []
    for idx in range(solid_count):
        solid_num = idx + 1
        stl_name = f"{instance_ref} - Solid {solid_num}.stl"
        stl_path = world_dir / stl_name
        if not stl_path.exists():
            return None  # incomplete set — fall through to worker
        results.append({
            "name": f"{instance_ref} - Solid {solid_num}",
            "stl_file": stl_name,
            "node_id": f"{node_id}:s{idx}",
            "size_bytes": stl_path.stat().st_size,
            "url": f"/outputs/stl/{run_id}/connpreview/{stl_name}",
        })

    logger.info(
        "conn_preview_reusing_stls",
        filename=filename,
        node_id=node_id,
        count=len(results),
    )
    return results


@router.post("/connections/preview/{filename}")
async def generate_preview_stl(
    filename: str,
    body: Dict[str, Any] = Body(default={}),
) -> Dict[str, Any]:
    """Generate per-solid STLs in world coordinates for 3D connection preview."""
    file_path = Path(settings.UPLOAD_DIR) / filename
    if not file_path.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error": "File not found", "filename": filename},
        )

    node_id: Optional[str] = body.get("node_id")
    if not node_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error": "node_id is required"},
        )

    # ── Fast path: reuse STLs already on disk (from analysis explode) ──
    cached_files = _find_existing_solid_stls(filename, node_id)
    if cached_files:
        return {
            "status": "completed",
            "files": cached_files,
            "node_id": node_id,
            "cached": True,
        }

    task_type = f"conn_preview:{node_id}"

    # Idempotent: return existing task if one matches
    existing_tasks = task_manager.get_tasks_for_file(filename, task_type=task_type)
    for task in existing_tasks:
        if task.status in (TaskStatus.PENDING, TaskStatus.RUNNING, TaskStatus.COMPLETED):
            return {
                "task_id": task.task_id,
                "status": task.status.value,
                "node_id": node_id,
                "existing": True,
            }

    output_dir = _stl_output_dir(filename)

    async def run_async(_progress_callback):
        cmd = [
            sys.executable, str(_STL_WORKER),
            "solids_world", str(file_path), str(output_dir), node_id,
        ]
        try:
            proc = await asyncio.create_subprocess_exec(
                *cmd,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
        except Exception as exc:
            raise RuntimeError(f"Failed to launch STL worker: {exc}")

        try:
            stdout_bytes, stderr_bytes = await asyncio.wait_for(
                proc.communicate(), timeout=float(_STL_TIMEOUT),
            )
        except asyncio.TimeoutError:
            proc.kill()
            await proc.wait()
            raise RuntimeError("Preview STL generation timed out")

        stdout = stdout_bytes.decode("utf-8", errors="replace")
        stderr = stderr_bytes.decode("utf-8", errors="replace").strip()

        if proc.returncode != 0:
            raise RuntimeError(
                f"STL worker failed (exit {proc.returncode}): {stderr}"
                if stderr
                else f"STL worker crashed (exit {proc.returncode})"
            )

        json_line = next(
            (line for line in stdout.splitlines() if line.strip().startswith("[")),
            stdout,
        )
        results = json.loads(json_line)

        # Build URL paths for each STL (connpreview/ subdirectory)
        run_id = filename[:8]
        for item in results:
            if item.get("stl_file"):
                item["url"] = f"/outputs/stl/{run_id}/connpreview/{item['stl_file']}"

        return results

    task_id = task_manager.submit_async(task_type, filename, run_async)
    logger.info(
        "conn_preview_submitted",
        filename=filename,
        task_id=task_id,
        node_id=node_id,
    )
    return {"task_id": task_id, "status": "pending", "node_id": node_id, "existing": False}


@router.post("/connections/preview-batch/{filename}")
async def generate_preview_stl_batch(
    filename: str,
    body: Dict[str, Any] = Body(default={}),
) -> Dict[str, Any]:
    """Generate per-solid STLs for multiple nodes in a single worker call.

    Much faster than calling /connections/preview N times because the STEP
    file is only parsed once.

    Body: {"node_ids": ["0:1:1:8:1", "0:1:1:8:2", ...]}
    """
    file_path = Path(settings.UPLOAD_DIR) / filename
    if not file_path.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error": "File not found", "filename": filename},
        )

    node_ids: List[str] = body.get("node_ids", [])
    if not node_ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error": "node_ids list is required"},
        )

    # Fast path: check which nodes already have cached STLs
    cached_files: List[Dict[str, Any]] = []
    uncached_ids: List[str] = []

    for nid in node_ids:
        files = _find_existing_solid_stls(filename, nid)
        if files:
            cached_files.extend(files)
        else:
            uncached_ids.append(nid)

    # If everything is cached, return immediately
    if not uncached_ids:
        return {
            "status": "completed",
            "files": cached_files,
            "cached": True,
        }

    # Check for existing batch preview task
    task_type = "conn_preview_batch"
    existing_tasks = task_manager.get_tasks_for_file(filename, task_type=task_type)
    for task in existing_tasks:
        if task.status in (TaskStatus.PENDING, TaskStatus.RUNNING):
            return {
                "task_id": task.task_id,
                "status": task.status.value,
                "existing": True,
                "cached_files": cached_files,
            }

    output_dir = _stl_output_dir(filename)

    async def run_async(_progress_callback):
        node_ids_arg = ",".join(uncached_ids)
        cmd = [
            sys.executable, str(_STL_WORKER),
            "solids_world_batch", str(file_path), str(output_dir), node_ids_arg,
        ]
        try:
            proc = await asyncio.create_subprocess_exec(
                *cmd,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
        except Exception as exc:
            raise RuntimeError(f"Failed to launch STL worker: {exc}")

        try:
            stdout_bytes, stderr_bytes = await asyncio.wait_for(
                proc.communicate(), timeout=float(_STL_TIMEOUT),
            )
        except asyncio.TimeoutError:
            proc.kill()
            await proc.wait()
            raise RuntimeError("Batch preview STL generation timed out")

        stdout = stdout_bytes.decode("utf-8", errors="replace")
        stderr = stderr_bytes.decode("utf-8", errors="replace").strip()

        if proc.returncode != 0:
            raise RuntimeError(
                f"STL worker failed (exit {proc.returncode}): {stderr}"
                if stderr
                else f"STL worker crashed (exit {proc.returncode})"
            )

        json_line = next(
            (line for line in stdout.splitlines() if line.strip().startswith("[")),
            stdout,
        )
        results = json.loads(json_line)

        # Build URL paths for each STL
        run_id = filename[:8]
        for item in results:
            if item.get("stl_file"):
                item["url"] = f"/outputs/stl/{run_id}/connpreview/{item['stl_file']}"

        # Combine cached + fresh results
        return cached_files + results

    task_id = task_manager.submit_async(task_type, filename, run_async)
    logger.info(
        "conn_preview_batch_submitted",
        filename=filename,
        task_id=task_id,
        cached=len(cached_files),
        uncached=len(uncached_ids),
    )
    return {
        "task_id": task_id,
        "status": "pending",
        "cached_files": cached_files,
        "existing": False,
    }


@router.get("/connections/preview-status/{task_id}")
async def get_preview_status(task_id: str) -> Dict[str, Any]:
    """Poll a preview STL generation task."""
    info = task_manager.get_task(task_id)
    if not info:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error": "Task not found", "task_id": task_id},
        )

    if info.status == TaskStatus.FAILED:
        return {"status": "failed", "error": info.error or "Preview generation failed"}

    if info.status in (TaskStatus.PENDING, TaskStatus.RUNNING):
        return {"status": info.status.value, "task_id": task_id}

    return {
        "status": "completed",
        "files": info.results or [],
        "filename": info.filename,
    }
